import streamlit as st

from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import os
import tempfile
import base64
from urllib.parse import unquote

# Load and encode logo
def get_base64_image(image_path):
    with open(image_path, "rb") as img_file:
        return base64.b64encode(img_file.read()).decode('utf-8')

logo_path = "logo.png"
logo_base64 = get_base64_image(logo_path) if os.path.exists(logo_path) else None

# Streamlit UI
st.markdown("""
    <h2 style='text-align: center; font-family: Bebas Neue Pro Expanded;'>MIRU Document Generator</h2>
    <hr style='margin-top: 0;'>
    <style>
        label, .stTextInput>div>div>input, .stTextArea textarea, .stNumberInput input {
            font-size: 16px;
            font-family: 'Bebas Neue Pro Expanded', sans-serif;
            color: #2E2E2E;
        }
        .stRadio label {
            font-size: 16px;
            font-family: 'Bebas Neue Pro Expanded', sans-serif;
        }
        .stDownloadButton>button {
            background-color: #1a1a1a;
            color: white;
            font-family: 'Bebas Neue Pro Expanded', sans-serif;
            letter-spacing: 1px;
        }
    </style>
""", unsafe_allow_html=True)

default_template_path = "default_invoice_template.xlsx"
default_wb = load_workbook(default_template_path) if os.path.exists(default_template_path) else None
default_ws = default_wb.active if default_wb else None

template_file = st.file_uploader("Upload Invoice Template (.xlsx)", type="xlsx")

query_params = st.query_params
client_name_q = unquote("".join(query_params.get("client", [""])))
billing_address_q = unquote("".join(query_params.get("billing", [""])))
delivery_address_q = unquote("".join(query_params.get("address", [""])))
qty_q = query_params.get("qty", [""])[0]
rate_q = query_params.get("rate", [""])[0]

if template_file or default_wb:
    wb = load_workbook(template_file) if template_file else default_wb
    ws = wb.active

    doc_type = st.selectbox("Document Type", ["Invoice", "Proforma Invoice", "Quotation"])

    st.markdown("## 👤 Client Details")
    client_name = st.text_input("Client Name", value=client_name_q)
    billing_address = st.text_area("Billing Address", value=billing_address_q)
    delivery_address = st.text_area("Delivery Address", value=delivery_address_q)
    invoice_date = st.date_input("Invoice Date", value=datetime.today())

    st.markdown("## 📄 Terms & Conditions")
    term_count = st.number_input("Number of terms", min_value=1, max_value=10, value=1)
    terms = [st.text_area(f"Term {i+1}", key=f"term_{i}") for i in range(term_count)]

    st.markdown("## 📦 Line Items")
    item_count = st.number_input("Number of line items", min_value=1, value=1)
    transport_included = st.radio("Transport Charges", ["Included", "Extra"], index=0)

    items = []
    for i in range(item_count):
        st.markdown(f"### Item {i+1}")
        hsn = st.text_input(f"HSN Code {i+1}", key=f"hsn_{i}")
        desc = st.text_input(f"Description {i+1}", key=f"desc_{i}")
        qty_default = float(qty_q) if i == 0 and qty_q else 0
        qty = st.number_input(f"Quantity {i+1}", key=f"qty_{i}", value=qty_default)
        unit = st.selectbox(f"Unit {i+1}", ["RFT", "SQFT", "SQM", "PC", "KG"], key=f"unit_{i}")
        rate_default = float(rate_q) if i == 0 and rate_q else 0
        rate = st.number_input(f"Rate {i+1}", key=f"rate_{i}", value=rate_default)
        items.append({"hsn": hsn, "desc": desc, "qty": qty, "unit": unit, "rate": rate})

    if st.button("Generate Document"):
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        font_table = Font(name="Bebas Neue Pro Expanded", color="595959")

        ws["A6"] = client_name
        ws["A6"].font = font_table
        ws["A7"] = billing_address
        ws["A7"].font = font_table
        ws["A14"] = delivery_address
        ws["A14"].font = font_table
        ws["I7"] = invoice_date.strftime('%d-%m-%Y')

        term_start_row = 34
        for i, term in enumerate(terms):
            row = term_start_row + i
            ws.merge_cells(f"A{row}:E{row}")
            ws[f"A{row}"] = f"{i+1}. {term.replace('\n', ' ').strip()}"
            ws[f"A{row}"].alignment = Alignment(wrap_text=False, vertical="top")

        for row in [21, 22]:
            for col in "ABCDEFGHI":
                ws[f"{col}{row}"].border = thin_border

        total = 0
        start_row = 23
        for i, item in enumerate(items):
            row = start_row + i
            if f"B{row}:D{row}" in [str(m) for m in ws.merged_cells.ranges]:
                ws.unmerge_cells(f"B{row}:D{row}")
            ws[f"A{row}"] = item["hsn"]
            ws.merge_cells(f"B{row}:D{row}")
            ws[f"B{row}"] = item["desc"].upper()
            ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
            ws.row_dimensions[row].height = 30 if len(item["desc"]) <= 50 else 45
            ws[f"B{row}"].font = Font(name="Bebas Neue Pro Expanded", color="595959")
            ws[f"F{row}"] = item["qty"]
            ws[f"G{row}"] = item["unit"].upper()
            ws[f"H{row}"] = item["rate"]
            amount = item["qty"] * item["rate"]
            ws[f"I{row}"] = f"₹{amount:,.2f}"
            total += amount
            for col in "ABCDFGHI":
                ws[f"{col}{row}"].border = thin_border
                ws[f"{col}{row}"].font = font_table

        grand_total = round(total * 1.18)
        summary_values = [f"₹{total:,.2f}", f"₹{total*0.09:,.2f}", f"₹{total*0.09:,.2f}", transport_included, f"₹{grand_total:,}"]

        html_path = os.path.join(os.path.dirname(__file__), "pdf.html")
        html_template = open(html_path, "r").read() if os.path.exists(html_path) else "<p><strong>HTML template missing.</strong></p>"
        
        # Inject logo
        logo_html = f"<img src='data:image/png;base64,{logo_base64}' style='height:80px;'>" if logo_base64 else "<strong>[Logo Missing]</strong>"
        html_filled = html_template
        html_filled = html_filled.replace("{{logo}}", logo_html)
        html_filled = html_filled.replace("{{document_type}}", doc_type)
        html_filled = html_filled.replace("{{recipient_name}}", client_name)
        html_filled = html_filled.replace("{{delivery_address}}", delivery_address.replace("\n", "<br>"))
        html_filled = html_filled.replace("{{invoice_date}}", invoice_date.strftime('%d-%m-%Y'))

        if items:
            item = items[0]
            html_filled = html_filled.replace("{{item_hsn}}", item['hsn'])
            html_filled = html_filled.replace("{{item_description}}", item['desc'])
            html_filled = html_filled.replace("{{item_qty}}", str(item['qty']))
            html_filled = html_filled.replace("{{item_unit}}", item['unit'])
            html_filled = html_filled.replace("{{item_rate}}", str(item['rate']))
            html_filled = html_filled.replace("{{item_amount}}", f"{item['qty'] * item['rate']:,.2f}")

        html_filled = html_filled.replace("{{subtotal}}", f"{total:,.2f}")
        html_filled = html_filled.replace("{{cgst}}", f"{total*0.09:,.2f}")
        html_filled = html_filled.replace("{{sgst}}", f"{total*0.09:,.2f}")
        html_filled = html_filled.replace("{{transportation}}", transport_included)
        html_filled = html_filled.replace("{{total_amount}}", f"{grand_total:,.2f}")
        html_filled = html_filled.replace("{{term_1}}", terms[0] if len(terms) > 0 else "")
        html_filled = html_filled.replace("{{term_2}}", terms[1] if len(terms) > 1 else "")
        html_filled = html_filled.replace("{{payment_terms}}", terms[2] if len(terms) > 2 else "")

        st.components.v1.html(html_filled, height=1000, scrolling=True)

        summary_labels = ["Subtotal", "CGST @ 9%", "SGST @ 9%", "Transport", "Total"]
        for i, (label, value) in enumerate(zip(summary_labels, summary_values)):
            row = start_row + len(items) + 1 + i
            ws[f"H{row}"] = label
            ws[f"I{row}"] = value
            ws[f"I{row}"].alignment = Alignment(horizontal="right")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
            tmp_excel.write(output.read())
            tmp_excel_path = tmp_excel.name

        doc_type_prefix = doc_type.replace(" ", "_").lower()
        filename = f"{doc_type_prefix}_{client_name.replace(' ', '_')}.xlsx"
        st.download_button("📥 Download Excel Document", data=output, file_name=filename)
